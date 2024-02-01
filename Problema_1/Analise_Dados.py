import pandas as pd
import re
from datetime import datetime
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# Variáveis e Listas
pasta = "Dados_Ext"
arquivos_csv = []
arquivos_csv_2 = []
arquivos_csv_3 = []
arquivos_csv_4 = []
n_amostra_lista = []
emp_sol_lista = []
end_lista = []
nome_sol_lista = []
ensaio_lista = []
id_cliente_lista = []
rotulada_lista = []
coletor_lista = []
data_amos_lista = []
data_ano_lista = []
tempoa_lista = []
data_lab_lista = []
data_RRA_lista = []
parametros_lista =[]
unidade_lista = []
diluicao_lista = []
lq_lista = []
resul_anal_lista = []
id_interno_texto = []
id_interno_lista = []


# Percorrer a pasta e encontrar arquivos CSV Area 1
for arquivo in os.listdir(pasta):
    if arquivo.endswith('.csv') and arquivo.startswith('arquivo_out_Area_1'):
        caminho_arquivo = os.path.join(pasta, arquivo)
        arquivos_csv.append(caminho_arquivo)

for arquivo_csv in arquivos_csv:
    # Área 1
    dados_1 = pd.read_csv(arquivo_csv)
    nome_coluna = dados_1.columns[0]
    padrao = r'\d+'
    numeros = re.findall(padrao, nome_coluna)
    n_amostra = ''.join(numeros)
    for _ in range(43):
        n_amostra_lista.append(n_amostra)
df_1= pd.DataFrame({"Número da amostra": n_amostra_lista})
df_1.to_csv('Dados_Analise/Dataframe_1.csv', index=False) 


# Percorrer a pasta e encontrar arquivos CSV Area 2
for arquivo in os.listdir(pasta):
    if arquivo.endswith('.csv') and arquivo.startswith('arquivo_out_Area_2'):
        caminho_arquivo = os.path.join(pasta, arquivo)
        arquivos_csv_2.append(caminho_arquivo) 

# Processar os arquivos CSV
for arquivo_csv in arquivos_csv_2:
    dados_2 = pd.read_csv(arquivo_csv)
    emp_sol = dados_2.columns[1]
    end = dados_2.iloc[0, 1]
    nome_sol = dados_2.iloc[1, 1]
    for _ in range(43):
        emp_sol_lista.append(emp_sol)
        end_lista.append(end)
        nome_sol_lista.append(nome_sol)
df_2= pd.DataFrame({"Empresa solicitante": emp_sol_lista,
                    "Endereço": end_lista,
                    "Nome do Solicitante": nome_sol_lista})
df_2.to_csv('Dados_Analise/Dataframe_2.csv', index=False) 


# Percorrer a pasta e encontrar arquivos CSV Area 3
for arquivo in os.listdir(pasta):
    if arquivo.endswith('.csv') and arquivo.startswith('arquivo_out_Area_3'):
        caminho_arquivo = os.path.join(pasta, arquivo)
        arquivos_csv_3.append(caminho_arquivo) 

# Processar os arquivos CSV
for arquivo_csv in arquivos_csv_3:
    dados_3 = pd.read_csv(arquivo_csv)
    ensaio = dados_3.columns[1]
    id_cliente = dados_3.iloc[0, 1].replace(" ", "")
    rotulada = dados_3.iloc[1, 1]
    coletor = dados_3.iloc[2, 1]
    #Datas
    data_str = dados_3.iloc[3, 1]
    data_obj = datetime.strptime(data_str, '%d/%m/%Y %H:%M')
    data_amos = int(data_obj.timestamp())
    data_ano = data_obj.strftime("%d/%m/%Y")
    tempo = data_obj.time()
    data_str_3 = dados_3.iloc[4, 2]
    data_RRA = datetime.strptime(data_str_3, "%d/%m/%Y")
    id_interno = f"{id_cliente}_{data_amos}"

    for _ in range(43):
        ensaio_lista.append(ensaio)
        id_cliente_lista.append(id_cliente)
        rotulada_lista.append(rotulada)
        coletor_lista.append(coletor)
        data_amos_lista.append(data_amos)
        data_ano_lista.append(data_ano)
        tempoa_lista.append(tempo)
        data_RRA_lista.append(data_RRA)
        id_interno_lista.append(id_interno)

df_3= pd.DataFrame({"Identificação do item de ensaio": ensaio_lista,
                    "Identificação do Cliente": id_cliente_lista,
                    "Amostra  Rotulada como": rotulada_lista,
                    'Coletor': coletor_lista,
                    'Data Amostragem': data_ano_lista,
                    'Hora amostragem' : tempoa_lista,
                    'Data RRA': data_RRA_lista,
                    'ID_interno': id_interno_lista
                    })
df_3.to_csv('Dados_Analise/Dataframe_3.csv', index=False)


# Percorrer a pasta e encontrar arquivos CSV Area 4
for arquivo in os.listdir(pasta):
    if arquivo.endswith('.csv') and arquivo.startswith('arquivo_out_Area_4'):
        caminho_arquivo = os.path.join(pasta, arquivo)
        arquivos_csv_4.append(caminho_arquivo) 

# Processar os arquivos CSV
for arquivo_csv in arquivos_csv_4:
    dados_4 = pd.read_csv(arquivo_csv)
    dados_4.columns = ['Parâmetros', 'Parâmetros_2', 'Unidade', 
                    'Diluição', 'Unnamed: 1','LQ / Faixa', 'Resultados analíticos', 
                    'Data  do Início', 'F1', 'F2']
    dados_4 = dados_4.drop(['Parâmetros_2', 'Data  do Início', 'F1', 'F2', 'Diluição'], axis=1)
    dados_4.columns = ['Parâmetros', 'Unidade', 'Diluição','LQ / Faixa', 'Resultados analíticos',]
    dados_4 = dados_4.drop(0).reset_index(drop=True)
    dados_4['Unidade'] = dados_4['Unidade'].str.replace('\u03BCg/m3', 'ug/m\u00B3')

    for valor in dados_4['Parâmetros']:
        parametros_lista.append(valor)
    for valor in dados_4['Unidade']:
        unidade_lista.append(valor)
    for valor in dados_4['Diluição']:
        diluicao_lista.append(valor)
    for valor in dados_4['LQ / Faixa']:
        lq_lista.append(valor)
    for valor in dados_4['Resultados analíticos']:
        resul_anal_lista.append(valor)

df_4= pd.DataFrame({"Parâmetros": parametros_lista,
                    "Unidade": unidade_lista,
                    "Diluição": diluicao_lista,
                    'LQ / Faixa': lq_lista,
                    'Resultados analíticos' : resul_anal_lista
                    })
df_4.to_csv('Dados_Analise/Dataframe_4.csv', index=False)



# Excel final (Visualização dos Dados)
wb = Workbook()
ws = wb.active

cabecalhos = ['Identificação interna', 'Nome da amostra', 
              'Data de coleta', 'Horário de coleta',
              'Parâmetro químico', 'Resultado', 'Unidade', 
              'Limite de Quantificação (LQ)', 't']
ws.append(cabecalhos)

#Dados
data_length = len(df_4['Parâmetros'])
for i in range(data_length):
    ws.append([
        id_interno_lista[i],
        id_cliente_lista[i],
        data_amos_lista[i],
        tempoa_lista[i],
        parametros_lista[i],
        resul_anal_lista[i],
        unidade_lista[i],
        500, 
        lq_lista[i]
        ])

# Formatação
ws.column_dimensions['I'].hidden = True # Ocultar a coluna 
for row in ws.iter_rows():
    for cell in row:
        #Centralizar
        cell.alignment = Alignment(horizontal='center', vertical='center')
        #Bordas
        cell.border = Border(left=Side(border_style='thin', color='000000'),
                             right=Side(border_style='thin', color='000000'),
                             top=Side(border_style='thin', color='000000'),
                             bottom=Side(border_style='thin', color='000000'))

# Auto ajuste largura
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        cell_length = len(str(cell.value))
        if cell_length > max_length:
            max_length = cell_length
        else:
            pass
    adjusted_width = (max_length + 2)
    ws.column_dimensions[column].width = adjusted_width

# Coluna resultado
for row_F, row_I in zip(ws.iter_rows(min_row=2, min_col=6, max_col=6), ws.iter_rows(min_row=2, min_col=9, max_col=9)):
    for cell_F, cell_I in zip(row_F, row_I):
        tamanho_F = len(str(cell_F.value)) if cell_F.value is not None else 0
        tamanho_I = len(str(cell_I.value)) if cell_I.value is not None else 0
        # Adiciona "<LQ" para valores numéricos ou aplica formatação em negrito
        if tamanho_F > tamanho_I:
            cell_F.value = "<LQ"
            cell_F.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid") 
        else:
            cell_F.font = Font(bold=True)
    

wb.save('Dataframe_Final.xlsx')




