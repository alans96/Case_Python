from openpyxl import load_workbook, Workbook


# Listas
valores_numeros_coluna1 = []
valores_numeros_coluna2 = []
valor_amb_aberto_C = []
valor_amb_aberto_NC = []
valor_amb_fechado_C = []
valor_amb_fechado_NC = []
sheet_2_coluna_1 = []
sheet_2_coluna_2 = []
sheet_2_coluna_3 = []
sheet_2_coluna_4 = []

# Abrir a planilha protegida por senha
workbook = load_workbook('Dados_Input/Material_Case_Ex2.xlsx', read_only=False, keep_vba=True)

# Desabilitar a proteção de estrutura (opcional)
"""for sheet in workbook:
    sheet.protection.sheet = False
workbook.save('Dados_Ext\EX_2_desprotegida.xlsx')
workbook.close()"""

# Tratamento de dados
sheet_1 = workbook['Avaliacao_Risco_Case']
sheet_2 = workbook['Valores_orientadores']

# Percorrer Coluna E e F e converter em 2 listas
for cell1, cell2 in zip(sheet_1['E'][7:], sheet_1['F'][7:]):
    if cell1.value is not None:
        # Tentar converter o valor da primeira célula para um número de ponto flutuante
        try:
            valor_convertido1 = float(cell1.value)
            valores_numeros_coluna1.append(valor_convertido1)
        except ValueError:
            # Se não for possível converter, adicionar 0 à lista
            valores_numeros_coluna1.append('0')

    # Verificar se o valor da segunda célula não é vazio
    if cell2.value is not None:
        try:
            valor_convertido2 = float(cell2.value)
            valores_numeros_coluna2.append(valor_convertido2)
        except ValueError:
            valores_numeros_coluna2.append('0')

# Separar os valores das colunas 1 e 2 em listas separadas de acordo com o índice
for i, (elemento1, elemento2) in enumerate(zip(valores_numeros_coluna1, valores_numeros_coluna2)):
    if i % 2 == 0:
        valor_amb_aberto_C.append(elemento1)
        valor_amb_fechado_C.append(elemento2)
    else:
        valor_amb_aberto_NC.append(elemento1)
        valor_amb_fechado_NC.append(elemento2)


# Percorrer Sheet_2 e adicionar em 4 listas
for row in sheet_2.iter_rows(values_only=True, min_row=2, max_row=sheet_2.max_row, min_col=1, max_col=4):
    sheet_2_coluna_1.append(row[0])
    sheet_2_coluna_2.append(row[1])
    sheet_2_coluna_3.append(row[2])
    sheet_2_coluna_4.append(row[3])

# Data base final
wb = Workbook()
ws = wb.active

cabecalhos = ['id','cas', 'contaminante',
                'vor', 'vor_valor',
                'abe_c', 'abe_nc',
                'fec_c', 'fec_nc',]
ws.append(cabecalhos)

# Dados
data_length = len(sheet_2_coluna_1)
for i in range(data_length):
    index = i+1
    ws.append([
        index,
        sheet_2_coluna_1[i],
        sheet_2_coluna_2[i],
        sheet_2_coluna_3[i],
        sheet_2_coluna_4[i],
        valor_amb_aberto_C[i],
        valor_amb_aberto_NC[i],
        valor_amb_fechado_C[i],
        valor_amb_fechado_NC[i]
        ])

wb.save('data_base.xlsx')

