import pandas as pd
import re
from datetime import datetime
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side


def centralizar_celulas(ws, inicio_linha, inicio_coluna, fim_linha, fim_coluna):
    for row in ws.iter_rows(min_row=inicio_linha, min_col=inicio_coluna, max_row=fim_linha, max_col=fim_coluna):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')



pasta = "data_base.xlsx"

Workbook = Workbook()
ws = Workbook.active
# Cabeçalho
ws.merge_cells(start_row=1, start_column=1, end_row=7, end_column=3)
ws.merge_cells(start_row=1, start_column=4, end_row=7, end_column=7)
ws.merge_cells(start_row=1, start_column=8, end_row=7, end_column=8)
ws.merge_cells(start_row=1, start_column=9, end_row=7, end_column=11)
ws.merge_cells(start_row=1, start_column=12, end_row=7, end_column=14)

ws.merge_cells(start_row=1, start_column=15, end_row=1, end_column=22)
ws.merge_cells(start_row=2, start_column=15, end_row=2, end_column=22)
ws.merge_cells(start_row=3, start_column=15, end_row=3, end_column=22)
ws.merge_cells(start_row=4, start_column=15, end_row=4, end_column=22)

ws.merge_cells(start_row=5, start_column=15, end_row=6, end_column=18)
ws.merge_cells(start_row=5, start_column=19, end_row=6, end_column=22)
ws.merge_cells(start_row=7, start_column=15, end_row=7, end_column=18)
ws.merge_cells(start_row=7, start_column=19, end_row=7, end_column=22)

ws['A1'] = "CAS"
ws['D1'] = "Substância Química de Interesse"
ws['H1'] = "Efeito"
ws['I1'] = "Concentração de solubilidade"
ws['L1'] = "VOR"
ws['O1'] = "ÁGUA SUBTERRÂNEA - TRABALHADOR COMERCIAL E INDUSTRIAL"
ws['O2'] = "UE-01A_Raso"
ws['O3'] = "Vias de exposição"
ws['O4'] = "Inalação"
ws['O5'] = "Ambiente Aberto"
ws['S5'] = "Ambiente Fechados"
ws['O7'] = "mg/L"
ws['S7'] = "mg/L"

# Alinhar Cabeçalho
centralizar_celulas(ws, 1, 1, 7, 3)
centralizar_celulas(ws, 1, 4, 7, 7)
centralizar_celulas(ws, 1, 8, 7, 8)
centralizar_celulas(ws, 1, 9, 7, 11)
centralizar_celulas(ws, 1, 12, 7, 14)
centralizar_celulas(ws, 1, 15, 1, 22)
centralizar_celulas(ws, 2, 15, 2, 22)
centralizar_celulas(ws, 3, 15, 3, 22)
centralizar_celulas(ws, 4, 15, 4, 22)
centralizar_celulas(ws, 5, 15, 6, 18)
centralizar_celulas(ws, 5, 19, 6, 22)
centralizar_celulas(ws, 7, 15, 7, 18)
centralizar_celulas(ws, 7, 19, 7, 22)

# Dados
for row in range(6, 30, 2):
    # Mesclar células com coluna 3
    for start_column in [1, 9]:
        ws.merge_cells(start_row=row, start_column=start_column, end_row=row+1, end_column=start_column+2)
# Mesclar células com coluna 4 
for row in range(6, 30,2):
    ws.merge_cells(start_row=row, start_column=4, end_row=row+1, end_column=7)
# Mesclar células com coluna 1 
for row in range(6, 30,2):
    ws.merge_cells(start_row=row, start_column=12, end_row=row+1, end_column=12)
# Mesclar células com coluna 2 
for row in range(6, 30, 2):
    for start_column in [13]:
        ws.merge_cells(start_row=row, start_column=start_column, end_row=row+1, end_column=start_column+1)


# Coluna Inalação
for row in range(7, 30):
    # Mesclar células para as colunas 15 e 19
    for start_column in [15, 19]:
        ws.merge_cells(start_row=row, start_column=start_column, end_row=row, end_column=start_column+1)

# Mesclar células para as colunas 17 e 21
for row in range(6, 30, 2):
    for start_column in [17, 21]:
        ws.merge_cells(start_row=row, start_column=start_column, end_row=row+1, end_column=start_column+1)

# Incerção de Dados
wb_dados = load_workbook(pasta)
ws_dados = wb_dados.active

# Incerção Coluna CAS
coluna_cas = [cell.value for cell in ws_dados["B"][1:]]
for index, value in enumerate(coluna_cas): # Coluna CAS
    ws.cell(row=8 + 2*index, column=1).value = value

# Incerção Coluna Contaminante
coluna_contaminante = [cell.value for cell in ws_dados["C"][1:]]
for index, value in enumerate(coluna_contaminante):
    ws.cell(row=8 + 2*index, column=4).value = value

# Incerção Coluna Efeito
efeito = ['C', 'NC']
for i in range(11):  # Repetir 11 vezes
    for index, value in enumerate(efeito):
        ws.cell(row=8 + 2*i + 1*index, column=8).value = value

# Incerção Coluna Concentração
concentracao = ['500']
for _ in range(11):  # Repetir 11 vezes
    for index, value in enumerate(concentracao):
        value = float(value)
        ws.cell(row=8 + 2*_ + 2*index, column=9).value = value
Workbook.save('relatorio.xlsx')

# Incerção Coluna vor_valor
# VOR Texto
concentracao = ['CETESB-ASUB_2021']
for _ in range(11):  # Repetir 11 vezes
    for index, value in enumerate(concentracao):
        ws.cell(row=8 + 2*_ + 2*index, column=13).value = value

# VOR Valor
coluna_vor = [cell.value for cell in ws_dados["E"][1:]]
coluna_vor_valor = [valor / 1000 for valor in coluna_vor]
for index, value in enumerate(coluna_vor_valor):
    ws.cell(row=8 + 2*index, column=12).value = value

# Incerção Coluna Ambiente Aberto
coluna_aber_C = [float(cell.value) for cell in ws_dados["F"][1:]]
coluna_aber_NC = [float(cell.value) for cell in ws_dados["G"][1:]]
amb_abert = list(zip(coluna_aber_C, coluna_aber_NC))

for i, (value_C, value_NC) in enumerate(amb_abert):
    ws.cell(row=8 + 2*i, column=15).value = value_C
    ws.cell(row=8 + 2*i + 1, column=15).value = value_NC

# Incerção Coluna Ambiente Fechado
coluna_fec_C = [float(cell.value) for cell in ws_dados["H"][1:]]
coluna_fec_NC = [float(cell.value) for cell in ws_dados["I"][1:]]
amb_fec = list(zip(coluna_fec_C, coluna_fec_NC))

for i, (value_C, value_NC) in enumerate(amb_fec):
    ws.cell(row=8 + 2*i, column=19).value = value_C
    ws.cell(row=8 + 2*i + 1, column=19).value = value_NC




# Analise do Dados

# Condiçõs para o Ambiente ABERTO   
for linha in range(1, 12):
    elemento1 = float(ws_dados["F"][linha].value)
    elemento2 = float(ws_dados["G"][linha].value)
    linha_ws = 6 + linha * 2 # Determina numero de iterações

    # Verificação 1
    menor_elemento = min(elemento1, elemento2)
    
    # Verificação 2
    ws.cell(row=linha_ws, column=17).value = menor_elemento
    if menor_elemento > 500:
        celula = ws.cell(row=linha_ws, column=17)
        celula.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    
    # Verificação 3
    valor_celula_12 = ws.cell(row=linha_ws, column=12).value
    if menor_elemento < valor_celula_12:
        ws.cell(row=linha_ws, column=17).value = valor_celula_12
        celula = ws.cell(row=linha_ws, column=17)
        ws.cell(row=linha_ws, column=17).font = Font(color="FFA500")


# Condiçõs para o Ambiente FECHADO
for linha in range(1, 12):
    elemento1 = float(ws_dados["h"][linha].value)
    elemento2 = float(ws_dados["i"][linha].value)
    linha_ws = 6 + linha * 2 # Determina numero de iterações

    # Verificação 1
    menor_elemento = min(elemento1, elemento2)

    # Verificação 2
    ws.cell(row=linha_ws, column=21).value = menor_elemento
    if menor_elemento > 500:
        celula = ws.cell(row=linha_ws, column=21)
        celula.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    # Verificação 3
    valor_celula_12 = ws.cell(row=linha_ws, column=12).value
    if menor_elemento < valor_celula_12:
        ws.cell(row=linha_ws, column=21).value = valor_celula_12
        celula = ws.cell(row=linha_ws, column=21)
        ws.cell(row=linha_ws, column=21).font = Font(color="FFA500")


Workbook.save('relatorio.xlsx')



